import csv
from datetime import timedelta
from pathlib import Path

from openpyxl import load_workbook

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.views.decorators.csrf import requires_csrf_token
from django.db.models import Q, Sum
from django.http import FileResponse, Http404, HttpResponse, HttpResponseForbidden, HttpResponseNotAllowed, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .audit import log_action
from .forms import (
    ActivityFundRequestForm,
    MemberForm,
    MemberImportForm,
    MemberNoteForm,
    MemberPortalAccessForm,
    MessageTemplateForm,
    PaymentCreateForm,
    PaymentForm,
    PublicPortalAccessForm,
    RefundForm,
    ReminderSettingForm,
)
from .models import (
    ActivityFundRequest,
    AuditLog,
    Member,
    MemberNote,
    MessageTemplate,
    NotificationLog,
    Payment,
    PaymentApprovalRequest,
    ReminderSetting,
)
from .notifications import send_member_reminder, send_payment_confirmation
from .roles import ROLE_PATRON, ROLE_TREASURER, capability_required, has_capability, user_roles

SMART_WATER_REPORT_NAME = "smart_water_metering_final_year_report_expanded.docx"


@requires_csrf_token
def csrf_failure(request, reason=""):
    return render(request, "members/csrf_failure.html", {"reason": reason}, status=403)


@login_required
def dashboard(request):
    if not has_capability(request.user, "view_dashboard"):
        if has_capability(request.user, "view_member_portal"):
            return redirect("member_portal")
        return HttpResponseForbidden("You do not have permission to access this page.")

    members = Member.objects.all()
    payments = Payment.objects.all()

    total_expected = sum(member.total_expected() for member in members)
    total_paid = payments.aggregate(total=Sum("amount"))["total"] or 0
    total_balance = sum(member.balance() for member in members)
    debtors = [member for member in members if member.is_debtor()]

    recent_payments = payments.select_related("member")[:5]
    recent_members = members.order_by("-created_at")[:5]
    failed_notifications = NotificationLog.objects.filter(status=NotificationLog.STATUS_FAILED).count()
    due_this_week = [member for member in members if member.next_due_date() <= timezone.localdate() + timedelta(days=7)]
    reminders_due_today = len(_reminders_due_on(timezone.localdate()))
    monthly_members = members.filter(plan=Member.PLAN_MONTHLY).count()
    four_month_members = members.filter(plan=Member.PLAN_FOUR_MONTHS).count()

    context = {
        "total_members": members.count(),
        "total_debtors": len(debtors),
        "total_expected": total_expected,
        "total_paid": total_paid,
        "total_balance": total_balance,
        "recent_payments": recent_payments,
        "recent_members": recent_members,
        "today": timezone.localdate(),
        "failed_notifications": failed_notifications,
        "due_this_week": len(due_this_week),
        "reminders_due_today": reminders_due_today,
        "monthly_members": monthly_members,
        "four_month_members": four_month_members,
    }
    return render(request, "members/dashboard.html", context)


@login_required
@capability_required("view_reports")
def smart_water_report(request):
    report_path = _smart_water_report_path()
    context = {
        "report_name": "Smart Water Metering Final Year Report",
        "file_name": SMART_WATER_REPORT_NAME,
        "file_size_kb": round(report_path.stat().st_size / 1024, 1) if report_path.exists() else None,
    }
    return render(request, "members/smart_water_report.html", context)


@login_required
@capability_required("view_reports")
def download_smart_water_report(request):
    report_path = _smart_water_report_path()
    if not report_path.exists():
        raise Http404("The project report file could not be found.")
    return FileResponse(
        report_path.open("rb"),
        as_attachment=True,
        filename="SMART_WATER_METERING_FINAL_YEAR_REPORT.docx",
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@login_required
@capability_required("view_members")
def member_list(request):
    query = request.GET.get("q", "").strip()
    plan = request.GET.get("plan", "").strip()
    members = Member.objects.all()

    if query:
        members = (
            members.filter(full_name__icontains=query)
            | members.filter(phone__icontains=query)
            | members.filter(member_id__icontains=query)
        )
    if plan in {Member.PLAN_MONTHLY, Member.PLAN_FOUR_MONTHS}:
        members = members.filter(plan=plan)

    return render(
        request,
        "members/member_list.html",
        {
            "members": members,
            "query": query,
            "plan": plan,
            "plan_choices": Member.PLAN_CHOICES,
            "monthly_count": Member.objects.filter(plan=Member.PLAN_MONTHLY).count(),
            "four_month_count": Member.objects.filter(plan=Member.PLAN_FOUR_MONTHS).count(),
        },
    )


@login_required
@capability_required("view_members")
def member_detail(request, pk):
    member = get_object_or_404(Member, pk=pk)
    payments = member.payments.all()
    notifications = member.notification_logs.all()[:5]
    notes = member.staff_notes.select_related("created_by").all()[:10]
    note_form = MemberNoteForm()
    breakdown = member.demand_breakdown()
    return render(
        request,
        "members/member_detail.html",
        {
            "member": member,
            "payments": payments,
            "breakdown": breakdown,
            "notifications": notifications,
            "notes": notes,
            "note_form": note_form,
        },
    )


@login_required
@capability_required("manage_members")
def member_create(request):
    form = MemberForm(request.POST or None)
    if form.is_valid():
        member = form.save()
        log_action(request, AuditLog.ACTION_CREATE, member, details="Registered a new member.")
        return redirect("member_detail", pk=member.pk)
    return render(request, "members/form.html", {"form": form, "title": "Register Member", "button": "Save Member"})


@login_required
@capability_required("manage_members")
def member_update(request, pk):
    member = get_object_or_404(Member, pk=pk)
    form = MemberForm(request.POST or None, instance=member)
    if form.is_valid():
        member = form.save()
        log_action(request, AuditLog.ACTION_UPDATE, member, details="Updated member details.")
        return redirect("member_detail", pk=member.pk)
    return render(request, "members/form.html", {"form": form, "title": "Edit Member", "button": "Update Member"})


@login_required
@capability_required("manage_member_portal_access")
def member_portal_access(request, pk):
    member = get_object_or_404(Member, pk=pk)
    form = MemberPortalAccessForm(request.POST or None, instance=member)
    if form.is_valid():
        member = form.save()
        log_action(request, AuditLog.ACTION_UPDATE, member, details="Updated member portal access.")
        messages.success(request, "Portal access updated.")
        return redirect("member_detail", pk=member.pk)
    return render(
        request,
        "members/form.html",
        {
            "form": form,
            "title": f"Portal Access - {member.full_name}",
            "button": "Save Portal Access",
            "cancel_url": "member_detail",
            "cancel_pk": member.pk,
        },
    )


@login_required
@capability_required("view_payments")
def payment_list(request):
    payments = Payment.objects.select_related("member").all()
    query = request.GET.get("q", "").strip()
    if query:
        payments = payments.filter(member__full_name__icontains=query) | payments.filter(received_by__icontains=query)
    return render(request, "members/payment_list.html", {"payments": payments, "query": query})


@login_required
@capability_required("record_payments")
def payment_create(request):
    received_by_name = request.user.get_full_name() or request.user.get_username()
    form = PaymentCreateForm(request.POST or None, received_by_name=received_by_name)
    member_id = request.GET.get("member")
    if member_id and not request.POST:
        form.fields["member"].initial = member_id

    if form.is_valid():
        payment = form.save(commit=False)
        payment.status = Payment.STATUS_COMPLETED
        payment.received_by = received_by_name
        payment.save()
        log_action(request, AuditLog.ACTION_CREATE, payment, details=f"Recorded payment for {payment.member.full_name}.")
        _, _, results = send_payment_confirmation(payment)
        sent = [result for result in results if result["status"] == NotificationLog.STATUS_SENT]
        failed = [result for result in results if result["status"] == NotificationLog.STATUS_FAILED]
        if sent:
            messages.success(request, f"Payment confirmation sent through {', '.join(result['channel'] for result in sent)}.")
        if failed:
            messages.warning(request, f"Payment was saved, but these confirmation channels failed: {', '.join(result['channel'] for result in failed)}.")
        if not results:
            messages.info(request, "Payment saved. No confirmation channels are enabled.")
        return redirect("payment_receipt", pk=payment.pk)

    return render(request, "members/form.html", {"form": form, "title": "Record Payment", "button": "Save Payment"})


@login_required
@capability_required("request_payment_actions")
def payment_update(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    form = PaymentForm(request.POST or None, instance=payment)
    if form.is_valid():
        if _requires_payment_approval(request):
            PaymentApprovalRequest.objects.create(
                payment=payment,
                requested_by=request.user,
                action=PaymentApprovalRequest.ACTION_EDIT,
                reason="Payment edit request.",
                proposed_data=_payment_form_data(form),
            )
            log_action(request, AuditLog.ACTION_UPDATE, payment, details="Requested approval to edit payment.")
            messages.info(request, "Payment edit request submitted for approval.")
        else:
            payment = form.save()
            log_action(request, AuditLog.ACTION_UPDATE, payment, details=f"Updated payment for {payment.member.full_name}.")
            messages.success(request, "Payment updated.")
        return redirect("payment_receipt", pk=payment.pk)
    return render(request, "members/form.html", {"form": form, "title": "Edit Payment", "button": "Update Payment"})


@login_required
@capability_required("request_payment_actions")
def payment_delete(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    member_pk = payment.member.pk
    if request.method == "POST":
        delete_reason = request.POST.get("reason", "").strip() or "Unintended payment recorded."
        if _requires_payment_approval(request):
            PaymentApprovalRequest.objects.create(
                payment=payment,
                requested_by=request.user,
                action=PaymentApprovalRequest.ACTION_DELETE,
                reason=delete_reason,
            )
            log_action(request, AuditLog.ACTION_DELETE, payment, details=f"Requested approval to delete unintended payment. Reason: {delete_reason}")
            messages.info(request, "Unintended payment delete request submitted for approval.")
        else:
            log_action(
                request,
                AuditLog.ACTION_DELETE,
                object_type="Payment",
                object_id=str(payment.pk),
                object_repr=str(payment),
                details=f"Deleted unintended payment for {payment.member.full_name}. Reason: {delete_reason}",
            )
            payment.delete()
            messages.success(request, "Unintended payment deleted.")
        return redirect("member_detail", pk=member_pk)
    return render(request, "members/confirm_delete.html", {"object": payment, "title": "Delete Unintended Payment"})


@login_required
@capability_required("request_payment_actions")
def payment_refund(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    form = RefundForm(request.POST or None, instance=payment)
    if form.is_valid():
        PaymentApprovalRequest.objects.create(
            payment=payment,
            requested_by=request.user,
            action=PaymentApprovalRequest.ACTION_REFUND,
            reason=form.cleaned_data["refund_reason"],
            proposed_data={
                "refunded_amount": form.cleaned_data["refunded_amount"],
                "refund_reason": form.cleaned_data["refund_reason"],
            },
        )
        log_action(request, AuditLog.ACTION_UPDATE, payment, details="Requested approval to refund payment.")
        messages.info(request, "Refund request submitted to Action Approvals.")
        return redirect("payment_approvals")
    return render(request, "members/form.html", {"form": form, "title": "Refund Payment", "button": "Save Refund"})


@login_required
@capability_required("view_members")
def debtors(request):
    members = [member for member in Member.objects.all() if member.is_debtor()]
    members.sort(key=lambda member: member.days_overdue(), reverse=True)
    total_demand = sum(member.balance() for member in members)
    return render(request, "members/debtors.html", {"members": members, "total_demand": total_demand})


@login_required
@capability_required("view_reports")
def reports(request):
    members = Member.objects.all()
    payments = Payment.objects.select_related("member").all()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)

    aging = {
        "Current": 0,
        "1-29 days": 0,
        "30-59 days": 0,
        "60-89 days": 0,
        "90+ days": 0,
    }
    for member in members:
        if member.balance() > 0:
            aging[member.overdue_bucket()] += member.balance()

    context = {
        "members": members,
        "payments": payments,
        "total_expected": sum(member.total_expected() for member in members),
        "total_paid": sum(payment.amount for payment in payments),
        "total_balance": sum(member.balance() for member in members),
        "aging": aging,
        "date_from": date_from,
        "date_to": date_to,
        "collections_by_method": _collections_by_method(payments),
    }
    return render(request, "members/reports.html", context)


@login_required
@capability_required("view_members")
def due_calendar(request):
    today = timezone.localdate()
    month = int(request.GET.get("month", today.month))
    year = int(request.GET.get("year", today.year))
    members = [
        member
        for member in Member.objects.filter(is_active=True)
        if member.next_due_date().month == month and member.next_due_date().year == year
    ]
    members.sort(key=lambda member: member.next_due_date())
    return render(
        request,
        "members/due_calendar.html",
        {
            "members": members,
            "month": month,
            "year": year,
            "total_due": sum(member.balance() for member in members if member.balance() > 0),
        },
    )


@login_required
@capability_required("send_reminders")
def bulk_reminders(request):
    MessageTemplate.ensure_defaults()
    selected_plan = request.GET.get("plan", "").strip()
    member_queryset = Member.objects.filter(is_active=True)
    if selected_plan in {Member.PLAN_MONTHLY, Member.PLAN_FOUR_MONTHS}:
        member_queryset = member_queryset.filter(plan=selected_plan)

    members = [member for member in member_queryset if member.balance() > 0]
    members.sort(key=lambda member: member.balance(), reverse=True)
    monthly_template = MessageTemplate.objects.get(kind=NotificationLog.KIND_MONTHLY_START)
    four_month_template = MessageTemplate.objects.get(kind=NotificationLog.KIND_FOUR_MONTH_REMINDER)
    preview_member = members[0] if members else None
    monthly_preview = monthly_template.render_body(preview_member) if preview_member else ""
    four_month_preview = four_month_template.render_body(preview_member) if preview_member else ""

    if request.method == "POST":
        selected_ids = request.POST.getlist("member_ids")
        kind = request.POST.get("kind", NotificationLog.KIND_MONTHLY_START)
        selected_members = [
            member
            for member in Member.objects.filter(pk__in=selected_ids, is_active=True)
            if member.balance() > 0
        ]
        skipped_count = len(selected_ids) - len(selected_members)
        sent_count = 0
        failed_count = 0

        for member in selected_members:
            _, _, results = send_member_reminder(member, kind, timezone.localdate(), force=True)
            sent_count += len([result for result in results if result["status"] == NotificationLog.STATUS_SENT])
            failed_count += len([result for result in results if result["status"] == NotificationLog.STATUS_FAILED])

        messages.success(request, f"Bulk reminders processed. Sent: {sent_count}. Failed: {failed_count}. Skipped no-balance/inactive members: {skipped_count}.")
        log_action(
            request,
            AuditLog.ACTION_BULK_REMINDER,
            object_type="Reminder",
            object_repr="Bulk reminder",
            details=f"Selected members: {len(selected_ids)}. Eligible debtors: {len(selected_members)}. Sent: {sent_count}. Failed: {failed_count}. Skipped: {skipped_count}.",
        )
        return redirect("notification_list")

    return render(
        request,
        "members/bulk_reminders.html",
        {
            "members": members,
            "selected_plan": selected_plan,
            "plan_choices": Member.PLAN_CHOICES,
            "monthly_template": monthly_template,
            "four_month_template": four_month_template,
            "preview_member": preview_member,
            "monthly_preview": monthly_preview,
            "four_month_preview": four_month_preview,
        },
    )


@login_required
@capability_required("view_payment_actions")
def payment_approvals(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    action = request.GET.get("action", "").strip()
    all_approvals = PaymentApprovalRequest.objects.all()
    status_counts = {
        PaymentApprovalRequest.STATUS_PENDING: all_approvals.filter(status=PaymentApprovalRequest.STATUS_PENDING).count(),
        PaymentApprovalRequest.STATUS_APPROVED: all_approvals.filter(status=PaymentApprovalRequest.STATUS_APPROVED).count(),
        PaymentApprovalRequest.STATUS_REJECTED: all_approvals.filter(status=PaymentApprovalRequest.STATUS_REJECTED).count(),
    }
    approvals = PaymentApprovalRequest.objects.select_related("payment", "payment__member", "requested_by").all()

    if status in {
        PaymentApprovalRequest.STATUS_PENDING,
        PaymentApprovalRequest.STATUS_APPROVED,
        PaymentApprovalRequest.STATUS_REJECTED,
    }:
        approvals = approvals.filter(status=status)

    if action in {
        PaymentApprovalRequest.ACTION_EDIT,
        PaymentApprovalRequest.ACTION_DELETE,
        PaymentApprovalRequest.ACTION_REFUND,
    }:
        approvals = approvals.filter(action=action)

    if query:
        approvals = approvals.filter(
            Q(payment__member__full_name__icontains=query)
            | Q(payment__member__member_id__icontains=query)
            | Q(requested_by__username__icontains=query)
            | Q(requested_by__first_name__icontains=query)
            | Q(requested_by__last_name__icontains=query)
            | Q(reason__icontains=query)
        )

    return render(
        request,
        "members/payment_approvals.html",
        {
            "approvals": approvals,
            "query": query,
            "status": status,
            "action": action,
            "status_choices": PaymentApprovalRequest.STATUS_CHOICES,
            "action_choices": PaymentApprovalRequest.ACTION_CHOICES,
            "status_counts": status_counts,
        },
    )


@login_required
@capability_required("approve_payment_actions")
def review_payment_approval(request, pk, decision):
    if not _can_approve_payments(request):
        return HttpResponseForbidden("You do not have permission to approve payment changes.")
    approval = get_object_or_404(PaymentApprovalRequest, pk=pk)
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    if decision == "approve":
        _apply_payment_approval(approval, request)
        approval.status = PaymentApprovalRequest.STATUS_APPROVED
        messages.success(request, "Action request approved.")
    else:
        approval.status = PaymentApprovalRequest.STATUS_REJECTED
        messages.info(request, "Action request rejected.")
    approval.reviewed_by = request.user
    approval.reviewed_at = timezone.now()
    approval.save()
    return redirect("payment_approvals")


@login_required
@capability_required("manage_roles")
def role_management(request):
    User = get_user_model()
    role_names = [ROLE_PATRON, ROLE_TREASURER]
    groups = {name: Group.objects.get_or_create(name=name)[0] for name in role_names}

    if request.method == "POST":
        user = get_object_or_404(User, pk=request.POST.get("user_id"))
        role = request.POST.get("role", "").strip()

        user.groups.remove(*groups.values())
        if role in groups:
            user.groups.add(groups[role])
            message = f"{user.get_username()} assigned as {role}."
        else:
            message = f"{user.get_username()} role cleared."

        log_action(request, AuditLog.ACTION_SETTINGS, object_type="User", object_id=str(user.pk), object_repr=user.get_username(), details=message)
        messages.success(request, message)
        return redirect("role_management")

    users = User.objects.prefetch_related("groups").order_by("username")
    user_rows = []
    for user in users:
        roles = user_roles(user)
        selected_role = ROLE_PATRON if ROLE_PATRON in roles else ROLE_TREASURER if ROLE_TREASURER in roles else ""
        user_rows.append(
            {
                "user": user,
                "roles": roles,
                "selected_role": selected_role,
            }
        )
    return render(
        request,
        "members/role_management.html",
        {
            "user_rows": user_rows,
            "role_choices": role_names,
        },
    )


@login_required
@capability_required("request_activity_funds")
def request_activity_fund(request):
    form = ActivityFundRequestForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        fund_request = form.save(commit=False)
        fund_request.requested_by = request.user
        fund_request.status = ActivityFundRequest.STATUS_PENDING
        fund_request.save()
        log_action(
            request,
            AuditLog.ACTION_CREATE,
            fund_request,
            details=f"Requested UGX {fund_request.amount:,} for club activity: {fund_request.activity_title}.",
        )
        messages.success(request, "Activity fund request submitted for approval.")
        return redirect("activity_fund_requests")

    return render(
        request,
        "members/request_activity_fund.html",
        {
            "form": form,
        },
    )


@login_required
@capability_required("view_activity_funds")
def activity_fund_requests(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    fund_requests = ActivityFundRequest.objects.select_related("requested_by", "reviewed_by").all()

    if status in {
        ActivityFundRequest.STATUS_PENDING,
        ActivityFundRequest.STATUS_APPROVED,
        ActivityFundRequest.STATUS_REJECTED,
    }:
        fund_requests = fund_requests.filter(status=status)

    if query:
        fund_requests = fund_requests.filter(
            Q(activity_title__icontains=query)
            | Q(description__icontains=query)
            | Q(requested_by__username__icontains=query)
            | Q(requested_by__first_name__icontains=query)
            | Q(requested_by__last_name__icontains=query)
            | Q(reviewed_by__username__icontains=query)
            | Q(reviewed_by__first_name__icontains=query)
            | Q(reviewed_by__last_name__icontains=query)
            | Q(review_note__icontains=query)
        )

    return render(
        request,
        "members/activity_fund_requests.html",
        {
            "fund_requests": fund_requests,
            "can_approve": _can_approve_fund_requests(request),
            "query": query,
            "status": status,
            "status_choices": ActivityFundRequest.STATUS_CHOICES,
        },
    )


@login_required
@capability_required("approve_activity_funds")
def review_activity_fund_request(request, pk, decision):
    if not _can_approve_fund_requests(request):
        return HttpResponseForbidden("You do not have permission to approve activity fund requests.")
    fund_request = get_object_or_404(ActivityFundRequest, pk=pk)
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    if fund_request.status != ActivityFundRequest.STATUS_PENDING:
        messages.info(request, "This activity fund request has already been reviewed.")
        return redirect("activity_fund_requests")

    if decision == "approve":
        fund_request.status = ActivityFundRequest.STATUS_APPROVED
        message = "Activity fund request approved."
    elif decision == "reject":
        fund_request.status = ActivityFundRequest.STATUS_REJECTED
        message = "Activity fund request rejected."
    else:
        return HttpResponseNotAllowed(["POST"])

    fund_request.reviewed_by = request.user
    fund_request.reviewed_at = timezone.now()
    fund_request.review_note = request.POST.get("review_note", "").strip()
    fund_request.save()
    log_action(
        request,
        AuditLog.ACTION_UPDATE,
        fund_request,
        details=f"{message} Amount: UGX {fund_request.amount:,}.",
    )
    messages.success(request, message)
    return redirect("activity_fund_requests")


@login_required
@capability_required("view_members")
def add_member_note(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    form = MemberNoteForm(request.POST)
    if form.is_valid():
        note = form.save(commit=False)
        note.member = member
        note.created_by = request.user
        note.save()
        log_action(request, AuditLog.ACTION_CREATE, note, details=f"Added note for {member.full_name}.")
        messages.success(request, "Member note added.")
    return redirect("member_detail", pk=member.pk)


@login_required
@capability_required("view_notifications")
def notification_list(request):
    notifications = NotificationLog.objects.select_related("member").all()[:200]
    return render(request, "members/notification_list.html", {"notifications": notifications})


@login_required
@capability_required("view_audit_trail")
def audit_trail(request):
    logs = AuditLog.objects.select_related("user").all()
    action = request.GET.get("action", "").strip()
    query = request.GET.get("q", "").strip()

    if action:
        logs = logs.filter(action=action)
    if query:
        logs = (
            logs.filter(object_repr__icontains=query)
            | logs.filter(object_type__icontains=query)
            | logs.filter(details__icontains=query)
            | logs.filter(user__username__icontains=query)
        )

    return render(
        request,
        "members/audit_trail.html",
        {
            "logs": logs[:300],
            "action": action,
            "query": query,
            "action_choices": AuditLog.ACTION_CHOICES,
        },
    )


@login_required
@capability_required("manage_reminder_settings")
def reminder_settings(request):
    MessageTemplate.ensure_defaults()
    setting = ReminderSetting.current()
    monthly_template = MessageTemplate.objects.get(kind=NotificationLog.KIND_MONTHLY_START)
    four_month_template = MessageTemplate.objects.get(kind=NotificationLog.KIND_FOUR_MONTH_REMINDER)
    payment_template = MessageTemplate.objects.get(kind=NotificationLog.KIND_PAYMENT_CONFIRMATION)

    setting_form = ReminderSettingForm(request.POST or None, instance=setting, prefix="settings")
    monthly_form = MessageTemplateForm(request.POST or None, instance=monthly_template, prefix="monthly")
    four_month_form = MessageTemplateForm(request.POST or None, instance=four_month_template, prefix="four")
    payment_form = MessageTemplateForm(request.POST or None, instance=payment_template, prefix="payment")

    if (
        request.method == "POST"
        and setting_form.is_valid()
        and monthly_form.is_valid()
        and four_month_form.is_valid()
        and payment_form.is_valid()
    ):
        setting_form.save()
        monthly_form.save()
        four_month_form.save()
        payment_form.save()
        log_action(request, AuditLog.ACTION_SETTINGS, object_type="ReminderSetting", object_repr="Reminder settings", details="Updated reminder channels and message templates.")
        messages.success(request, "Reminder settings saved.")
        return redirect("reminder_settings")

    return render(
        request,
        "members/reminder_settings.html",
        {
            "setting_form": setting_form,
            "monthly_form": monthly_form,
            "four_month_form": four_month_form,
            "payment_form": payment_form,
        },
    )


@login_required
@capability_required("send_reminders")
def send_member_reminder_view(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    member = get_object_or_404(Member, pk=pk)
    kind = request.POST.get("kind", NotificationLog.KIND_MONTHLY_START)
    if kind not in {NotificationLog.KIND_MONTHLY_START, NotificationLog.KIND_FOUR_MONTH_REMINDER}:
        kind = NotificationLog.KIND_MONTHLY_START

    _, _, results = send_member_reminder(member, kind, timezone.localdate(), force=True)
    sent = [result for result in results if result["status"] == NotificationLog.STATUS_SENT]
    failed = [result for result in results if result["status"] == NotificationLog.STATUS_FAILED]

    if sent:
        messages.success(request, f"Reminder sent through {', '.join(result['channel'] for result in sent)}.")
    if failed:
        messages.warning(request, f"Some channels failed: {', '.join(result['channel'] for result in failed)}.")
    if not results:
        messages.info(request, "No reminder channels are enabled.")

    log_action(
        request,
        AuditLog.ACTION_SEND_REMINDER,
        member,
        details=f"Reminder type: {kind}. Sent: {len(sent)}. Failed: {len(failed)}.",
    )
    return redirect("member_detail", pk=member.pk)


@login_required
@capability_required("view_members")
def member_statement(request, pk):
    member = get_object_or_404(Member, pk=pk)
    payments = member.payments.all()
    return render(
        request,
        "members/member_statement.html",
        {"member": member, "payments": payments, "breakdown": member.demand_breakdown(), "today": timezone.localdate()},
    )


@login_required
@capability_required("view_payments")
def payment_receipt(request, pk):
    payment = get_object_or_404(Payment.objects.select_related("member"), pk=pk)
    return render(request, "members/payment_receipt.html", {"payment": payment, "today": timezone.localdate()})


@login_required
def member_portal(request):
    member = getattr(request.user, "member_profile", None)
    if not member:
        messages.info(request, "This login account is not linked to a member profile yet.")
        return render(request, "members/member_portal_unlinked.html")
    return render(
        request,
        "members/member_portal.html",
        {
            "member": member,
            "payments": member.payments.all(),
            "breakdown": member.demand_breakdown(),
            "notifications": member.notification_logs.all()[:10],
        },
    )


def public_member_portal(request):
    member_id = request.session.get("public_portal_member_id")
    if member_id:
        member = get_object_or_404(Member, pk=member_id, is_active=True)
        return render(
            request,
            "members/public_member_portal.html",
            {
                "member": member,
                "payments": member.payments.all(),
                "breakdown": member.demand_breakdown(),
            },
        )

    form = PublicPortalAccessForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        entered_member_id = form.cleaned_data["member_id"].strip().upper()
        member = Member.objects.filter(member_id__iexact=entered_member_id, is_active=True).first()
        if member and _phones_match(member.phone, form.cleaned_data["phone"]):
            request.session["public_portal_member_id"] = member.pk
            return redirect("public_member_portal")
        messages.warning(request, "No member matched that ID and phone number.")

    return render(request, "members/public_portal_access.html", {"form": form})


def public_member_portal_logout(request):
    request.session.pop("public_portal_member_id", None)
    messages.info(request, "Member portal session closed.")
    return redirect("public_member_portal")


@login_required
@capability_required("import_members")
def import_members(request):
    form = MemberImportForm(request.POST or None, request.FILES or None)
    imported = 0
    if request.method == "POST" and form.is_valid():
        uploaded_file = form.cleaned_data["file"]
        try:
            rows = _member_import_rows(uploaded_file)
        except ValueError as exc:
            messages.warning(request, str(exc))
            return redirect("import_members")

        for row in rows:
            name = (row.get("full_name") or row.get("name") or "").strip()
            phone = (row.get("phone") or "").strip()
            if not name:
                continue
            member = _find_member_by_phone(phone)
            defaults = {
                "full_name": name,
                "phone": phone,
                "email": (row.get("email") or "").strip(),
                "plan": (row.get("plan") or Member.PLAN_MONTHLY).strip() or Member.PLAN_MONTHLY,
            }
            if member:
                for field, value in defaults.items():
                    setattr(member, field, value)
                member.save()
            else:
                Member.objects.create(**defaults)
            imported += 1
        log_action(request, AuditLog.ACTION_CREATE, object_type="Member", object_repr="Member import", details=f"Imported or updated {imported} members using telephone number matching.")
        messages.success(request, f"Imported or updated {imported} members.")
        return redirect("member_list")
    return render(request, "members/import_members.html", {"form": form})


@login_required
@capability_required("export_reports")
def export_members_csv(request):
    log_action(request, AuditLog.ACTION_EXPORT, object_type="Member", object_repr="members.csv", details="Exported members CSV.")
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="members.csv"'
    writer = csv.writer(response)
    writer.writerow(["Member ID", "Name", "Phone", "Email", "Joined", "Plan", "Expected", "Paid", "Balance", "Next Due", "Status"])
    for member in Member.objects.all():
        writer.writerow(
            [
                member.member_id,
                member.full_name,
                member.phone,
                member.email,
                member.joined_date,
                member.get_plan_display(),
                member.total_expected(),
                member.total_paid(),
                member.balance(),
                member.next_due_date(),
                member.status(),
            ]
        )
    return response


@login_required
@capability_required("view_members")
def api_members(request):
    data = [
        {
            "id": member.pk,
            "member_id": member.member_id,
            "full_name": member.full_name,
            "phone": member.phone,
            "email": member.email,
            "plan": member.plan,
            "balance": member.balance(),
            "next_due_date": member.next_due_date().isoformat(),
            "status": member.status(),
        }
        for member in Member.objects.all()
    ]
    return JsonResponse({"members": data})


@login_required
@capability_required("view_payments")
def api_payments(request):
    data = [
        {
            "id": payment.pk,
            "member_id": payment.member_id,
            "member": payment.member.full_name,
            "amount": payment.amount,
            "payment_date": payment.payment_date.isoformat(),
            "method": payment.method,
            "reference": payment.reference,
            "status": payment.status,
        }
        for payment in Payment.objects.select_related("member").all()
    ]
    return JsonResponse({"payments": data})


@login_required
@capability_required("export_reports")
def export_payments_csv(request):
    log_action(request, AuditLog.ACTION_EXPORT, object_type="Payment", object_repr="payments.csv", details="Exported payments CSV.")
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="payments.csv"'
    writer = csv.writer(response)
    writer.writerow(["Member", "Payment Date", "Amount", "Received By", "Note", "Recorded At"])
    for payment in Payment.objects.select_related("member").all():
        writer.writerow(
            [
                payment.member.full_name,
                payment.payment_date,
                payment.amount,
                payment.received_by,
                payment.note,
                payment.created_at,
            ]
        )
    return response


def _smart_water_report_path():
    return Path(settings.BASE_DIR) / "static" / "reports" / SMART_WATER_REPORT_NAME


def _reminders_due_on(run_date):
    setting = ReminderSetting.current()
    members = [member for member in Member.objects.filter(is_active=True) if member.balance() > 0]
    if run_date.day == setting.monthly_reminder_day:
        return members

    due_members = []
    for member in members:
        months_due = member.months_due_on(run_date)
        previous_months_due = member.months_due_on(run_date - timedelta(days=1))
        if member.plan == Member.PLAN_FOUR_MONTHS and months_due != previous_months_due and months_due > 0 and months_due % 4 == 0:
            due_members.append(member)
    return due_members


def _collections_by_method(payments):
    summary = {}
    for payment in payments:
        summary[payment.get_method_display()] = summary.get(payment.get_method_display(), 0) + payment.amount
    return summary


def _phones_match(saved_phone, entered_phone):
    saved = "".join(char for char in saved_phone if char.isdigit())
    entered = "".join(char for char in entered_phone if char.isdigit())
    return bool(saved and entered and (saved.endswith(entered) or entered.endswith(saved)))


def _find_member_by_phone(phone):
    cleaned = "".join(char for char in phone if char.isdigit())
    if not cleaned:
        return None

    matches = []
    for member in Member.objects.exclude(phone=""):
        member_phone = "".join(char for char in member.phone if char.isdigit())
        if member_phone and (member_phone.endswith(cleaned) or cleaned.endswith(member_phone)):
            matches.append(member)

    if len(matches) == 1:
        return matches[0]
    return None


def _member_import_rows(uploaded_file):
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        uploaded_file.seek(0)
        lines = uploaded_file.read().decode("utf-8-sig").splitlines()
        return list(csv.DictReader(lines))

    if filename.endswith(".xlsx"):
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
        data = []
        for values in rows[1:]:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else ""
                row[header] = "" if value is None else str(value).strip()
            data.append(row)
        return data

    raise ValueError("Upload a CSV or XLSX Excel file.")


def _requires_payment_approval(request):
    return not has_capability(request.user, "approve_payment_actions")


def _can_approve_payments(request):
    return has_capability(request.user, "approve_payment_actions")


def _can_approve_fund_requests(request):
    return has_capability(request.user, "approve_activity_funds")


def _payment_form_data(form):
    data = {}
    for field, value in form.cleaned_data.items():
        if field == "member":
            data["member_id"] = value.pk
        elif hasattr(value, "isoformat"):
            data[field] = value.isoformat()
        else:
            data[field] = value
    return data


def _apply_payment_approval(approval, request):
    payment = approval.payment
    if approval.action == PaymentApprovalRequest.ACTION_EDIT:
        data = approval.proposed_data
        payment.member_id = data["member_id"]
        payment.amount = data["amount"]
        payment.payment_date = timezone.datetime.fromisoformat(data["payment_date"]).date()
        payment.method = data["method"]
        payment.reference = data["reference"]
        payment.provider = data["provider"]
        payment.external_transaction_id = data["external_transaction_id"]
        payment.status = data["status"]
        payment.received_by = data["received_by"]
        payment.note = data["note"]
        payment.save()
        log_action(request, AuditLog.ACTION_UPDATE, payment, details=f"Approved and applied payment edit request #{approval.pk}.")
    elif approval.action == PaymentApprovalRequest.ACTION_REFUND:
        payment.refunded_amount = approval.proposed_data.get("refunded_amount") or 0
        payment.refund_reason = approval.proposed_data.get("refund_reason") or ""
        payment.status = Payment.STATUS_REFUNDED
        payment.save()
        log_action(request, AuditLog.ACTION_UPDATE, payment, details=f"Approved refund request #{approval.pk}.")
    elif approval.action == PaymentApprovalRequest.ACTION_DELETE:
        log_action(request, AuditLog.ACTION_DELETE, payment, details=f"Approved and deleted payment request #{approval.pk}.")
        payment.delete()
